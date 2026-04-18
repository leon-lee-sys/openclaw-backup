#!/usr/bin/env python3
"""
通用数据可视化图表生成器
支持：人力资源、财务、资产管理、项目管理等数据
"""

import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import sys
import os

def style_header(ws, row=1, color="4472C4"):
    """样式化表头"""
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = thin_border

def auto_width(ws):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_bar_chart(ws, data_range, title, x_label, y_label, chart_position):
    """创建柱状图"""
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.y_axis.title = y_label
    chart.x_axis.title = x_label
    chart.style = 10
    
    data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=len(data_range))
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(data_range))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.legend = None
    ws.add_chart(chart, chart_position)

def create_line_chart(ws, data_range, title, x_label, y_label, chart_position):
    """创建折线图"""
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_label
    chart.x_axis.title = x_label
    chart.style = 10
    
    data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=len(data_range))
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(data_range))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.smooth = True
    chart.legend = None
    ws.add_chart(chart, chart_position)

def create_pie_chart(ws, data_range, title, chart_position):
    """创建饼图"""
    chart = PieChart()
    chart.title = title
    chart.style = 10
    
    data = Reference(ws, min_col=2, min_row=2, max_row=len(data_range))
    labels = Reference(ws, min_col=1, min_row=2, max_row=len(data_range))
    chart.add_data(data)
    chart.set_labels(labels)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showCatName = True
    ws.add_chart(chart, chart_position)

# ========== 人力资源模板 ==========
def generate_hr_dashboard(output_path="HR_数据分析看板.xlsx"):
    """生成人力资源数据分析看板"""
    wb = openpyxl.Workbook()
    
    # Sheet1: 员工结构分析
    ws1 = wb.active
    ws1.title = "员工结构"
    
    # 部门人数数据
    dept_data = [
        ["部门", "人数", "占比%"],
        ["技术部", 45, 30],
        ["销售部", 35, 23.3],
        ["市场部", 20, 13.3],
        ["财务部", 15, 10],
        ["人力资源", 12, 8],
        ["行政部", 18, 12],
        ["运营部", 5, 3.3],
    ]
    for row in dept_data:
        ws1.append(row)
    style_header(ws1)
    auto_width(ws1)
    
    # 学历分布
    ws2 = wb.create_sheet("学历分布")
    edu_data = [
        ["学历", "人数"],
        ["博士", 5],
        ["硕士", 35],
        ["本科", 80],
        ["大专", 25],
        ["其他", 5],
    ]
    for row in edu_data:
        ws2.append(row)
    style_header(ws2, color="70AD47")
    auto_width(ws2)
    
    # 司龄分布
    ws3 = wb.create_sheet("司龄分析")
    tenure_data = [
        ["司龄区间", "人数"],
        ["1年以下", 30],
        ["1-3年", 45],
        ["3-5年", 35],
        ["5-10年", 25],
        ["10年以上", 15],
    ]
    for row in tenure_data:
        ws3.append(row)
    style_header(ws3, color="ED7D31")
    auto_width(ws3)
    
    # 离职率趋势
    ws4 = wb.create_sheet("离职趋势")
    attrition_data = [
        ["月份", "离职率%"],
        ["1月", 2.1],
        ["2月", 1.8],
        ["3月", 2.5],
        ["4月", 3.2],
        ["5月", 2.8],
        ["6月", 3.5],
    ]
    for row in attrition_data:
        ws4.append(row)
    style_header(ws4, color="7030A0")
    auto_width(ws4)
    
    wb.save(output_path)
    print(f"✅ 人力资源看板已生成: {output_path}")
    return output_path

# ========== 财务管理模板 ==========
def generate_finance_dashboard(output_path="财务_数据分析看板.xlsx"):
    """生成财务数据分析看板"""
    wb = openpyxl.Workbook()
    
    # 收入分析
    ws1 = wb.active
    ws1.title = "收入分析"
    revenue_data = [
        ["月份", "收入(万)", "成本(万)", "利润(万)"],
        ["1月", 580, 420, 160],
        ["2月", 620, 450, 170],
        ["3月", 710, 510, 200],
        ["4月", 680, 490, 190],
        ["5月", 750, 540, 210],
        ["6月", 820, 580, 240],
    ]
    for row in revenue_data:
        ws1.append(row)
    style_header(ws1)
    auto_width(ws1)
    
    # 费用分析
    ws2 = wb.create_sheet("费用分析")
    expense_data = [
        ["费用类型", "金额(万)"],
        ["人力成本", 320],
        ["运营成本", 180],
        ["营销费用", 150],
        ["研发费用", 200],
        ["管理费用", 90],
        ["其他费用", 60],
    ]
    for row in expense_data:
        ws2.append(row)
    style_header(ws2, color="C00000")
    auto_width(ws2)
    
    # 现金流分析
    ws3 = wb.create_sheet("现金流")
    cashflow_data = [
        ["月份", "经营现金流", "投资现金流", "融资现金流"],
        ["Q1", 180, -50, 20],
        ["Q2", 220, -80, 0],
        ["Q3", 250, -60, -30],
        ["Q4", 300, -100, 50],
    ]
    for row in cashflow_data:
        ws3.append(row)
    style_header(ws3, color="4472C4")
    auto_width(ws3)
    
    wb.save(output_path)
    print(f"✅ 财务看板已生成: {output_path}")
    return output_path

# ========== 资产管理模板 ==========
def generate_asset_dashboard(output_path="资产_数据分析看板.xlsx"):
    """生成资产管理数据分析看板"""
    wb = openpyxl.Workbook()
    
    # 资产分类
    ws1 = wb.active
    ws1.title = "资产分类"
    asset_data = [
        ["资产类型", "原值(万)", "净值(万)", "折旧率%"],
        ["土地", 5000, 5000, 0],
        ["房屋建筑", 8000, 6400, 20],
        ["生产设备", 3500, 2100, 40],
        ["办公设备", 500, 200, 60],
        ["车辆", 800, 400, 50],
        ["软件系统", 300, 90, 70],
    ]
    for row in asset_data:
        ws1.append(row)
    style_header(ws1, color="548235")
    auto_width(ws1)
    
    # 资产分布
    ws2 = wb.create_sheet("资产分布")
    asset_dist = [
        ["区域", "资产值(万)"],
        ["华东区", 4500],
        ["华南区", 3200],
        ["华北区", 3800],
        ["西部地区", 2100],
        ["海外", 1500],
    ]
    for row in asset_dist:
        ws2.append(row)
    style_header(ws2, color="7030A0")
    auto_width(ws2)
    
    # 资产维护
    ws3 = wb.create_sheet("维护记录")
    maint_data = [
        ["月份", "维护次数", "维护费用(万)"],
        ["1月", 12, 3.5],
        ["2月", 15, 4.2],
        ["3月", 18, 5.1],
        ["4月", 14, 3.8],
        ["5月", 16, 4.5],
        ["6月", 20, 5.8],
    ]
    for row in maint_data:
        ws3.append(row)
    style_header(ws3, color="ED7D31")
    auto_width(ws3)
    
    wb.save(output_path)
    print(f"✅ 资产看板已生成: {output_path}")
    return output_path

# ========== 项目管理模板 ==========
def generate_project_dashboard(output_path="项目_数据分析看板.xlsx"):
    """生成项目管理数据分析看板"""
    wb = openpyxl.Workbook()
    
    # 项目进度
    ws1 = wb.active
    ws1.title = "项目进度"
    project_data = [
        ["项目名称", "计划进度%", "实际进度%", "偏差%"],
        ["数字化转型", 75, 72, -3],
        ["新工厂建设", 45, 48, 3],
        ["ERP升级", 60, 58, -2],
        ["人才招聘计划", 90, 88, -2],
        ["市场扩展", 30, 35, 5],
        ["研发中心", 55, 50, -5],
    ]
    for row in project_data:
        ws1.append(row)
    style_header(ws1, color="4472C4")
    auto_width(ws1)
    
    # 项目预算
    ws2 = wb.create_sheet("预算执行")
    budget_data = [
        ["项目", "预算(万)", "实际(万)", "执行率%"],
        ["数字化转型", 500, 420, 84],
        ["新工厂建设", 2000, 1050, 52.5],
        ["ERP升级", 300, 285, 95],
        ["人才招聘计划", 150, 145, 96.7],
        ["市场扩展", 800, 320, 40],
        ["研发中心", 600, 380, 63.3],
    ]
    for row in budget_data:
        ws2.append(row)
    style_header(ws2, color="70AD47")
    auto_width(ws2)
    
    # 风险分布
    ws3 = wb.create_sheet("风险分布")
    risk_data = [
        ["风险等级", "数量", "占比%"],
        ["高风险", 5, 16.7],
        ["中风险", 15, 50],
        ["低风险", 10, 33.3],
    ]
    for row in risk_data:
        ws3.append(row)
    style_header(ws3, color="C00000")
    auto_width(ws3)
    
    # 资源利用率
    ws4 = wb.create_sheet("资源利用")
    resource_data = [
        ["部门", "计划工时", "实际工时", "利用率%"],
        ["技术部", 2000, 1850, 92.5],
        ["销售部", 1500, 1620, 108],
        ["市场部", 800, 720, 90],
        ["财务部", 400, 380, 95],
        ["人力资源", 300, 295, 98.3],
    ]
    for row in resource_data:
        ws4.append(row)
    style_header(ws4, color="7030A0")
    auto_width(ws4)
    
    wb.save(output_path)
    print(f"✅ 项目管理看板已生成: {output_path}")
    return output_path

if __name__ == "__main__":
    print("📊 通用数据可视化图表生成器")
    print("=" * 40)
    
    output_dir = os.path.expanduser("~/Desktop/小燕子成果文件库/")
    os.makedirs(output_dir, exist_ok=True)
    
    # 生成各类型看板
    generate_hr_dashboard(os.path.join(output_dir, "HR_数据分析看板.xlsx"))
    generate_finance_dashboard(os.path.join(output_dir, "财务_数据分析看板.xlsx"))
    generate_asset_dashboard(os.path.join(output_dir, "资产_数据分析看板.xlsx"))
    generate_project_dashboard(os.path.join(output_dir, "项目_数据分析看板.xlsx"))
    
    print("=" * 40)
    print("🎉 全部看板生成完成！")
