#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "国学经典24章精读计划"

# 列宽设置
col_widths = {
    'A': 6,   # 序号
    'B': 18,  # 书名
    'C': 14,  # 作者
    'D': 14,  # 学习开始月份
    'E': 14,  # 学习结束月份
    'F': 12,  # 学习时长(月)
    'G': 40,  # 章节/内容概要
    'H': 10,  # 难度等级
    'I': 30,  # 计划备注
}

for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# 标题行样式
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 列标题
headers = ["序号", "书名", "作者", "学习开始月份", "学习结束月份", "学习时长(月)", "章节/内容概要", "难度等级", "计划备注"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

ws.row_dimensions[1].height = 36

# 数据填充
data = [
    # 2个月的书（8本）
    [1,  "《论语》",           "孔子及其弟子",    "2026.06", "2026.07", 2, "学而、为政、八佾、里仁等二十篇",             "★★★★☆", "儒家首要经典，建议精读细悟"],
    [2,  "《孟子》",           "孟轲",            "2026.08", "2026.09", 2, "梁惠王、公孙丑、滕文公等七篇",               "★★★★☆", "性善论核心，体悟仁政思想"],
    [3,  "《庄子·内七篇》",   "庄周",            "2026.10", "2026.11", 2, "逍遥游、齐物论、养生主等内七篇",             "★★★★★", "道家精髓，想象力丰富，需反复研读"],
    [4,  "《春秋左传》精选",  "左丘明",          "2026.12", "2027.01", 2, "郑伯克段、晋公子重耳、烛之武退秦师等",       "★★★★☆", "历史散文典范，叙事精彩"],
    [5,  "《史记》列传精选",  "司马迁",          "2027.02", "2027.03", 2, "项羽本纪、廉颇蔺相如、李将军列传等",        "★★★★☆", "史家之绝唱，文字优美"],
    [6,  "《资治通鉴》精选",  "司马光",          "2027.04", "2027.05", 2, "战国至五代重大政治事件精选",                "★★★★☆", "以史为鉴，识人用人之道"],
    [7,  "《礼记》精选",      "戴圣",            "2027.06", "2027.07", 2, "大学、中庸原文及礼运、檀弓等篇",            "★★★★☆", "儒家礼制与社会伦理"],
    [8,  "《汉书》精选",      "班固",            "2027.08", "2027.09", 2, "高帝纪、文帝纪、苏武传、东方朔传等",        "★★★★☆", "史学名著，了解汉代社会"],
    # 1个月的书（10本）
    [9,  "《大学》",           "曾子（孔伋）",   "2027.10", "2027.10", 1, "明明德、亲民、止于至善三纲领",              "★★★☆☆", "四书之一，儒学入门"],
    [10, "《中庸》",           "子思",           "2027.11", "2027.11", 1, "中、和、天命、率性等三十三章",              "★★★★☆", "四书之一，哲学性强"],
    [11, "《荀子》精选",       "荀况",           "2027.12", "2027.12", 1, "劝学、修身、非相、性恶等篇",                "★★★★☆", "性恶论，与孟子对照阅读"],
    [12, "《诗经》精选",       "佚名",           "2028.01", "2028.01", 1, "风、雅、颂精选篇目，含关雎、蒹葭等",        "★★★☆☆", "中国诗歌源头，吟诵体会"],
    [13, "《楚辞》精选",       "屈原等",         "2028.02", "2028.02", 1, "离骚、天问、九歌、招魂等",                 "★★★★☆", "浪漫主义文学巅峰"],
    [14, "《唐诗三百首》精选", "孙洙（蘅塘退士）","2028.03", "2028.03", 1, "五言律诗、七言律诗、五言绝句等代表作品",   "★★★☆☆", "诗歌入门必读，反复吟诵"],
    [15, "《古文观止》精选",   "吴楚材、吴调侯", "2028.04", "2028.04", 1, "左传、国语、史记、唐宋八大家散文精选",     "★★★★☆", "古文精选本，文以载道"],
    [16, "《易经》",           "佚名",           "2028.05", "2028.05", 1, "上经、下经、六十四卦卦辞与爻辞",          "★★★★★", "群经之首，义理与象数并重"],
    [17, "《传习录》",         "王阳明",         "2028.06", "2028.06", 1, "徐爱录、陆澄录、薛侃录等",                  "★★★★☆", "心学入门，知行合一"],
    [18, "《菜根谭》",         "洪应明",         "2028.07", "2028.07", 1, "修身、劝学、处世、人生等格言集",          "★★★☆☆", "处世智慧，淡然从容"],
    # 半个月的书（6本）
    [19, "《孝经》",           "孔伋",           "2028.08.1", "2028.08.15", 0.5, "开宗明义、天子、诸侯、卿大夫、士、庶人六章", "★★☆☆☆", "儒家伦理，短小精悍"],
    [20, "《孙子兵法》",       "孙武",           "2028.08.16", "2028.08.31", 0.5, "计篇、作战、谋攻、军形、兵势等十三篇",       "★★★★☆", "兵家圣典，管理借鉴"],
    [21, "《鬼谷子》",         "王诩",           "2028.09.1", "2028.09.15", 0.5, "捭阖、反应、内揵、抵巇等篇",                 "★★★★☆", "纵横家之祖，谋略之源"],
    [22, "《围炉夜话》",       "王永彬",         "2028.09.16", "2028.09.30", 0.5, "立身、处世、读书、治家格言一百三十余则",    "★★★☆☆", "儒家处世哲学，温暖人心"],
    [23, "《呻吟语》精选",     "吕坤",           "2028.10.1", "2028.10.15", 0.5, "内经、性命、存心、伦理、修身等精选",        "★★★★☆", "明代儒学精华"],
    [24, "《了凡四训》",       "袁了凡",         "2028.10.16", "2028.10.31", 0.5, "立命之学、改过之法、积善之方、谦德之效",    "★★★☆☆", "修身励志，因果智慧"],
]

# 数据行样式
row_fills = [
    PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
    PatternFill(start_color="EEF3FA", end_color="EEF3FA", fill_type="solid"),
]

data_font = Font(name="微软雅黑", size=10)
data_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

for row_idx, row_data in enumerate(data, 2):
    fill = row_fills[(row_idx) % 2]
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.border = thin_border
        cell.fill = fill
        if col_idx == 7 or col_idx == 9:  # 章节概要和计划备注左对齐
            cell.alignment = data_align_left
        else:
            cell.alignment = data_align
    ws.row_dimensions[row_idx].height = 32

# 保存
output_path = "/Users/mac/.openclaw/workspace/courses/国学经典24章精读_2年计划.xlsx"
wb.save(output_path)
print(f"✅ Excel文件已成功生成：{output_path}")